
from openpyxl import load_workbook

wb = load_workbook('../datas/daily_gabia_20171127.xlsx')
ws = wb.active

print(ws['A1'].value)
print(ws['E3'].value)

A = ws['E4':'E10']

for k in A:
    for m in k:
        print(m.value +2)

wb.save('../datas/daily_gabia_20171127_edit.xlsx')