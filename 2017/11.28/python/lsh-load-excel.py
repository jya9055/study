from openpyxl import load_workbook

wb = load_workbook('../datas/daily_gabia_20171127.xlsx') # 엑셀 열기
ws = wb.active # 현재 워크시크 가져오기

print('A1 = ' + ws['A1'].value) # A1 셀 내용 가져오기

A = ws['A4':'L4'] # A4:L4 셀 내용 가져오기

for t in A: # A에는 A4:L4 튜플이 저장됨 ( (), () )
    print('A4:L4 =')
    for c in t: # t는 튜플 ( )
        print(c.value)