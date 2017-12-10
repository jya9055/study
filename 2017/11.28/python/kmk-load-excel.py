from openpyxl import load_workbook

wb = load_workbook('../datas/daily_gabia_20171127.xlsx')
ws = wb.active

print('A1 = ' + ws['A1'].value)

A = ws['A4':'L4']

for k in A:
    print('A4:L4 =')
    for m in k:
        print(m.value)
