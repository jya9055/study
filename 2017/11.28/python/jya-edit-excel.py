# 환불 필드 E4부터 E10까지 +2

from openpyxl import load_workbook

wb = load_workbook('../datas/daily_gabia_20171127.xlsx')
ws = wb.active


# 환불 데이터 가져오기
for r in ws.rows:
    row_idx = r[4].row
    k = r[4].value
    if k in [2, 0, 12, 21, 8, 18]: 
        # 위 조건이 없을 경우, E1, E2, E3이 숫자가 아니어서 오류가 남 [LSH]: column_index_from_string 라는 함수로 E -> 4인 숫자형 인덱스로 변경 가능합니다.
        # if k in int 처럼 정수면 더하라는 조건을 추가하고 싶었으나, 못 찾아서 위처럼 처리
        p = k + 2

        # 2 더한 값 넣기
        ws.cell(row = row_idx, column=5).value = p
        # 근데 r[4]랑 column=5는 같은 의미 아닌가요??? [LSH]: r[4]는 각행의 4번째(E) 셀 즉 E1, E2, E3 을 나타냅니다. 그러나 E[4]는 <Cell ..> 와 같은 openpyxl에서 제공하는 자료형이고 column=5는 숫자 자료형입니다. 의미상으론 E(5번째)셀을 가리킨다는 것은 동일하나 엄연히 다릅니다.

    # [LSH]: 그냥 안쓰면 됩니다.
    # else: #else일 때 그냥 넘어가는 걸 어떻게 쓰나요??? return만 쓰면 빠져나온다고 했는데 outside function 오류남
    #     print("ddd")


wb.save('../datas/jya-edit.xlsx')
