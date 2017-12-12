# Let do it!
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

wb = load_workbook('../datas/daily_gabia_20171127.xlsx') # 엑셀 열기
ws = wb.active # 현재 워크시크 가져오기

# 1. 이중 for문이 불필요하다는 사실 인지. ws.cell은 행,열만 알면되는데 첫번째 for 문에서 다 알 수 있기 때문에
# 2. 상수로 선언했던 `column_idx = 5`를 변수로 처리할 방법 찾음. `E`라는 문자열 인덱스를 숫자형 인데스로 변환하는 column_index_from_string를 찾음
# 3. 내부 로직에 +2 라는 상수만 남음 +2 라는 상수도 변수 처리하면 순수 함수 생성 가능
# 4. 인자값 2개(셀 범위, 증가값)을 받는 함수 생성
# 5. 해당 함수의 내부 로직은 모두 변수화 됨 내부 상수 없음

def edit(cell, append):
    for t in cell: # ( (), () )
        row_idx = t[0].row # 현재 행 index
        column_idx = column_index_from_string(t[0].column)
        row_value = t[0].value
        ws.cell(row=row_idx, column=column_idx).value = row_value + append

E = ws['E4':'E10'] 

edit(E, 3)

wb.save('../datas/lsh-edit.xlsx') # 엑셀 파일 저장
wb.close()