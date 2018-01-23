from openpyxl import load_workbook

wb = load_workbook('../datas/daily_gabia_20171127.xlsx')
ws = wb.active

print('A1 = ' + ws['A1'].value)

A = ws['A4':'L4']

for t in A:
    print('A4:L4 =')
    for c in t:
        print(c.value)