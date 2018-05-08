from openpyxl import Workbook
from openpyxl.styles import Alignment
import sys
sys.path.append('C:/Users/조양아/Documents/GitHub/study/2018/04.10/python')
from jya_gAPIclass import GapiClass
# 이 파일과 동일한 경로에 config.py 추가 안하면 오류가 나는데, 모듈 불러올 때 config.py는 같이 안불러오나봐요??

wb = Workbook()
ws = wb.active

instance = GapiClass()
field_name_list = ['번호', '아이디', '이름']
id_list = ['planning_d', 'test1gabia', 'jya9055']
number = 0

# 필드명 입력

while number < len(field_name_list):
    field = field_name_list[number]
    number += 1
    ws.cell(row=1, column=number).value = field
    font = ws.cell(row=1, column=number)
    font.alignment = Alignment(horizontal='center')

number = 0

# 리스트 입력 

while number < len(id_list):
    id = id_list[number]
    number += 1
    ws.cell(row=number+1, column=1).value = number
    ws.cell(row=number+1, column=2).value = id
    ws.cell(row=number+1, column=3).value = instance.getMember(id)
    font = ws.cell(row=number+1, column=1)
    font.alignment = Alignment(horizontal='center')



wb.save('../datas/hanname-excel.xlsx')